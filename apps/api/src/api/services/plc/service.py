from __future__ import annotations

import csv
from datetime import datetime, timezone
import json
from pathlib import Path
import re
from typing import Any, Iterable, Literal

from openpyxl import load_workbook
from sqlalchemy import Select, select
from sqlalchemy.orm import Session

from api.config import get_settings
from api.models import PLCTestCaseRecord, PLCTestSuiteRecord
from api.services.jobs import to_iso
from api.services.plc.contracts import (
    PLCTestCaseModel,
    PLCTestSuiteDefinitionModel,
    PLCTestSuiteDetailModel,
    PLCTestSuiteSummaryModel,
)
from api.services.plc.persistence import (
    get_testcase_models_for_suite,
    resolve_plc_target,
)
from api.services.plc.profiles import (
    attach_execution_profile,
    ensure_execution_profiles,
    list_execution_profile_models,
)


class PLCImportError(RuntimeError):
    pass


def validate_plc_target(session: Session, *, target_key: str) -> dict[str, Any]:
    try:
        target = resolve_plc_target(session, target_key=target_key)
    except ValueError as exc:
        raise PLCImportError(str(exc)) from exc
    if target is None:
        raise PLCImportError(f"PLC target '{target_key}' was not found")
    if not bool(target.get("is_active")):
        raise PLCImportError(f"PLC target '{target_key}' is inactive")

    configured_executor_mode = get_settings().plc_executor_mode
    target_executor_mode = str(target.get("executor_mode") or "").strip().lower()
    if target_executor_mode and target_executor_mode != configured_executor_mode:
        raise PLCImportError(
            "PLC target "
            f"'{target_key}' requires executor mode '{target_executor_mode}', "
            f"but API is configured for '{configured_executor_mode}'"
        )
    return target


REQUIRED_COLUMNS = {
    "instruction_name",
    "input_values",
    "expected_outputs",
    "input_type",
    "output_type",
}
OPTIONAL_COLUMNS = {
    "description",
    "tags",
    "suite",
    "memory_profile_key",
    "timeout_ms",
    "expected_outcome",
    "case_key",
}


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _normalize_text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def _try_json_loads(value: str) -> Any:
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return value


def _coerce_sequence(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if isinstance(value, tuple):
        return list(value)
    if isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            return []
        parsed = _try_json_loads(normalized)
        if isinstance(parsed, list):
            return parsed
        if ";" in normalized:
            return [item.strip() for item in normalized.split(";") if item.strip()]
        if "," in normalized:
            return [item.strip() for item in normalized.split(",") if item.strip()]
        return [parsed]
    if value is None:
        return []
    return [value]


def _coerce_expected_outputs(value: Any) -> list[Any]:
    outputs = _coerce_sequence(value)
    if outputs and len(outputs) == 1 and isinstance(outputs[0], list):
        return outputs[0]
    return outputs


def _coerce_input_vectors(value: Any) -> list[list[Any]]:
    parsed = _coerce_sequence(value)
    if not parsed:
        return []
    if all(not isinstance(item, list) for item in parsed):
        return [parsed]
    vectors: list[list[Any]] = []
    for item in parsed:
        if isinstance(item, list):
            vectors.append(item)
        else:
            vectors.append([item])
    return vectors


def _coerce_tags(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            return []
        parsed = _try_json_loads(normalized)
        if isinstance(parsed, list):
            return [str(item).strip() for item in parsed if str(item).strip()]
        return [item.strip() for item in normalized.split(",") if item.strip()]
    return []


def _coerce_timeout(value: Any) -> int:
    if value in (None, ""):
        return 3000
    return max(1, int(value))


def _coerce_expected_outcome(value: Any) -> Literal["pass", "fail"]:
    normalized = str(value or "pass").strip().lower()
    return "fail" if normalized == "fail" else "pass"


def _read_csv_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(text.splitlines())
    return [dict(row) for row in reader]


def _read_xlsx_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    from io import BytesIO

    workbook = load_workbook(
        filename=BytesIO(file_bytes), read_only=True, data_only=True
    )
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_header(value) for value in rows[0]]
    normalized_rows: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        if raw_row is None:
            continue
        row_dict = {
            headers[index]: raw_row[index]
            for index in range(min(len(headers), len(raw_row)))
            if headers[index]
        }
        normalized_rows.append(row_dict)
    return normalized_rows


def _normalize_import_rows(
    raw_rows: Iterable[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str]]:
    cases: list[dict[str, Any]] = []
    warnings: list[str] = []
    for row_number, raw_row in enumerate(raw_rows, start=2):
        row = {
            _normalize_header(key): value
            for key, value in raw_row.items()
            if key is not None
        }
        if not any(value not in (None, "") for value in row.values()):
            continue
        missing = sorted(
            column
            for column in REQUIRED_COLUMNS
            if column not in row or row[column] in (None, "")
        )
        if missing:
            warnings.append(
                f"row {row_number}: missing required columns {', '.join(missing)}"
            )
            continue
        instruction_name = _normalize_text(row.get("instruction_name"))
        input_type = _normalize_text(row.get("input_type"))
        output_type = _normalize_text(row.get("output_type"))
        if not instruction_name or not input_type or not output_type:
            warnings.append(f"row {row_number}: required text fields must not be empty")
            continue

        input_vectors = _coerce_input_vectors(row.get("input_values"))
        expected_outputs = _coerce_expected_outputs(row.get("expected_outputs"))
        if not input_vectors:
            warnings.append(
                f"row {row_number}: input_values produced no runnable cases"
            )
            continue
        if not expected_outputs:
            warnings.append(
                f"row {row_number}: expected_outputs produced no runnable cases"
            )
            continue
        if len(expected_outputs) not in {1, len(input_vectors)}:
            warnings.append(
                f"row {row_number}: expected_outputs count {len(expected_outputs)} does not match input_values count {len(input_vectors)}"
            )
            continue

        tags = _coerce_tags(row.get("tags"))
        base_case_key = _normalize_text(row.get("case_key")) or instruction_name.upper()
        timeout_ms = _coerce_timeout(row.get("timeout_ms"))
        expected_outcome = _coerce_expected_outcome(row.get("expected_outcome"))
        for index, input_vector in enumerate(input_vectors, start=1):
            expected_output = (
                expected_outputs[index - 1]
                if len(expected_outputs) > 1
                else expected_outputs[0]
            )
            cases.append(
                {
                    "case_key": f"{base_case_key}_{index:03d}",
                    "instruction_name": instruction_name,
                    "input_type": input_type,
                    "output_type": output_type,
                    "input_vector_json": input_vector,
                    "expected_output_json": expected_output,
                    "expected_outputs_json": expected_outputs,
                    "memory_profile_key": _normalize_text(
                        row.get("memory_profile_key")
                    ),
                    "description": _normalize_text(row.get("description")),
                    "tags": tags,
                    "timeout_ms": timeout_ms,
                    "source_row_number": row_number,
                    "source_case_index": index - 1,
                    "expected_outcome": expected_outcome,
                }
            )
    return cases, warnings


def _next_suite_id(session: Session) -> str:
    next_id = 1
    for existing_id in session.scalars(select(PLCTestSuiteRecord.id)).all():
        match = re.search(r"(\d+)$", str(existing_id))
        if match is None:
            continue
        next_id = max(next_id, int(match.group(1)) + 1)
    return f"plc-suite-{next_id}"


def _suite_definition_with_ids(
    suite_id: str, cases: list[dict[str, Any]], warnings: list[str]
) -> PLCTestSuiteDefinitionModel:
    materialized_cases = [
        attach_execution_profile(
            PLCTestCaseModel(
                id=f"{suite_id}::{case['case_key']}",
                **case,
            )
        )
        for case in cases
    ]
    return PLCTestSuiteDefinitionModel(cases=materialized_cases, warnings=warnings)


def _serialize_suite(record: PLCTestSuiteRecord) -> PLCTestSuiteSummaryModel:
    return PLCTestSuiteSummaryModel(
        id=record.id,
        title=record.title,
        source_filename=record.source_filename,
        source_format=record.source_format,
        case_count=record.case_count,
        created_at=to_iso(record.created_at),
        updated_at=to_iso(record.updated_at),
    )


def _case_model_to_record(
    suite_id: str, case: PLCTestCaseModel, *, is_active: bool = True
) -> PLCTestCaseRecord:
    return PLCTestCaseRecord(
        id=case.id,
        suite_id=suite_id,
        testcase_key=case.id,
        case_key=case.case_key,
        instruction_name=case.instruction_name,
        input_type=case.input_type,
        output_type=case.output_type,
        input_vector_json=case.input_vector_json,
        expected_output_json=case.expected_output_json,
        expected_outputs_json=case.expected_outputs_json,
        expected_outcome=case.expected_outcome,
        description=case.description,
        tags_json=case.tags,
        memory_profile_key=case.memory_profile_key,
        execution_profile_key=case.execution_profile_key,
        timeout_ms=case.timeout_ms,
        source_row_number=case.source_row_number,
        source_case_index=case.source_case_index,
        is_active=is_active,
        updated_at=datetime.now(timezone.utc),
    )


def _serialize_testcase_record(
    record: PLCTestCaseRecord,
    *,
    suite_title: str | None = None,
    execution_profile: dict[str, Any] | None = None,
    case_source: str = "relational",
) -> dict[str, Any]:
    return {
        "id": record.id,
        "suite_id": record.suite_id,
        "suite_title": suite_title,
        "case_source": case_source,
        "testcase_key": record.testcase_key,
        "case_key": record.case_key,
        "instruction_name": record.instruction_name,
        "input_type": record.input_type,
        "output_type": record.output_type,
        "input_vector_json": record.input_vector_json,
        "expected_output_json": record.expected_output_json,
        "expected_outputs_json": record.expected_outputs_json,
        "memory_profile_key": record.memory_profile_key,
        "execution_profile_key": record.execution_profile_key,
        "execution_profile": execution_profile,
        "description": record.description,
        "tags": record.tags_json,
        "timeout_ms": record.timeout_ms,
        "source_row_number": record.source_row_number,
        "source_case_index": record.source_case_index,
        "expected_outcome": record.expected_outcome,
        "is_active": record.is_active,
        "created_at": to_iso(record.created_at),
        "updated_at": to_iso(record.updated_at),
    }


def _record_to_case_model(record: PLCTestCaseRecord) -> PLCTestCaseModel:
    expected_outcome: Literal["pass", "fail"] = (
        "fail" if record.expected_outcome == "fail" else "pass"
    )
    return PLCTestCaseModel(
        id=record.id,
        case_key=record.case_key,
        instruction_name=record.instruction_name,
        input_type=record.input_type,
        output_type=record.output_type,
        input_vector_json=record.input_vector_json,
        expected_output_json=record.expected_output_json,
        expected_outputs_json=record.expected_outputs_json,
        memory_profile_key=record.memory_profile_key,
        execution_profile_key=record.execution_profile_key,
        description=record.description,
        tags=record.tags_json,
        timeout_ms=record.timeout_ms,
        source_row_number=record.source_row_number,
        source_case_index=record.source_case_index,
        expected_outcome=expected_outcome,
    )


def ensure_plc_testcase_records(
    session: Session,
    *,
    suite_id: str,
    cases: list[PLCTestCaseModel],
) -> None:
    ensure_execution_profiles(session, cases=cases)
    existing_case_ids = set(
        session.scalars(
            select(PLCTestCaseRecord.id).where(PLCTestCaseRecord.suite_id == suite_id)
        ).all()
    )
    missing_records = [
        _case_model_to_record(suite_id, case)
        for case in cases
        if case.id not in existing_case_ids
    ]
    if missing_records:
        session.add_all(missing_records)
        session.flush()


def import_plc_suite(
    session: Session,
    *,
    filename: str,
    file_bytes: bytes,
    title: str | None,
) -> tuple[PLCTestSuiteDetailModel, int, int]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        raw_rows = _read_csv_rows(file_bytes)
        source_format = "csv"
    elif suffix == ".xlsx":
        raw_rows = _read_xlsx_rows(file_bytes)
        source_format = "xlsx"
    else:
        raise PLCImportError("only .csv and .xlsx files are supported")
    if not raw_rows:
        raise PLCImportError("import file contained no data rows")

    cases, warnings = _normalize_import_rows(raw_rows)
    if not cases:
        raise PLCImportError("no valid PLC testcases were found in the import file")

    suite_id = _next_suite_id(session)
    suite_definition = _suite_definition_with_ids(suite_id, cases, warnings)
    suite = PLCTestSuiteRecord(
        id=suite_id,
        title=(title or Path(filename).stem).strip() or Path(filename).stem,
        source_filename=filename,
        source_format=source_format,
        case_count=len(suite_definition.cases),
        definition_json=suite_definition.model_dump(mode="json"),
        updated_at=datetime.now(timezone.utc),
    )
    session.add(suite)
    ensure_execution_profiles(session, cases=suite_definition.cases)
    session.add_all(
        [_case_model_to_record(suite.id, case) for case in suite_definition.cases]
    )
    session.commit()
    session.refresh(suite)
    return (
        PLCTestSuiteDetailModel(
            **_serialize_suite(suite).model_dump(),
            definition_json=PLCTestSuiteDefinitionModel.model_validate(
                suite.definition_json
            ),
        ),
        len(suite_definition.cases),
        len(warnings),
    )


def list_plc_suites(session: Session) -> list[PLCTestSuiteSummaryModel]:
    suites = session.scalars(
        select(PLCTestSuiteRecord).order_by(
            PLCTestSuiteRecord.created_at.desc(), PLCTestSuiteRecord.id.desc()
        )
    ).all()
    return [_serialize_suite(suite) for suite in suites]


def get_plc_suite_detail(
    session: Session, suite_id: str
) -> PLCTestSuiteDetailModel | None:
    suite = session.get(PLCTestSuiteRecord, suite_id)
    if suite is None:
        return None
    return PLCTestSuiteDetailModel(
        **_serialize_suite(suite).model_dump(),
        definition_json=PLCTestSuiteDefinitionModel.model_validate(
            suite.definition_json
        ),
    )


def flatten_cases(
    session: Session,
    *,
    instruction_name: str | None = None,
    input_type: str | None = None,
    suite_id: str | None = None,
    tag: str | None = None,
) -> list[dict[str, Any]]:
    requested_suite_ids: list[str] | None = [suite_id] if suite_id is not None else None
    case_stmt: Select[tuple[PLCTestCaseRecord]] = select(PLCTestCaseRecord).where(
        PLCTestCaseRecord.is_active.is_(True)
    )
    if suite_id is not None:
        case_stmt = case_stmt.where(PLCTestCaseRecord.suite_id == suite_id)
    if instruction_name is not None:
        case_stmt = case_stmt.where(
            PLCTestCaseRecord.instruction_name == instruction_name
        )
    if input_type is not None:
        case_stmt = case_stmt.where(PLCTestCaseRecord.input_type == input_type)

    cases = session.scalars(
        case_stmt.order_by(
            PLCTestCaseRecord.suite_id.asc(),
            PLCTestCaseRecord.source_row_number.asc(),
            PLCTestCaseRecord.source_case_index.asc(),
            PLCTestCaseRecord.id.asc(),
        )
    ).all()
    if cases:
        suite_ids = sorted({case.suite_id for case in cases})
        execution_profiles = list_execution_profile_models(
            session,
            keys=[case.execution_profile_key or "" for case in cases],
        )
        suite_map = {
            suite.id: suite
            for suite in session.scalars(
                select(PLCTestSuiteRecord).where(PLCTestSuiteRecord.id.in_(suite_ids))
            ).all()
        }
        items = []
        for case in cases:
            if tag and tag not in case.tags_json:
                continue
            suite = suite_map.get(case.suite_id)
            items.append(
                _serialize_testcase_record(
                    case,
                    suite_title=suite.title if suite is not None else None,
                    execution_profile=(
                        execution_profiles[case.execution_profile_key].model_dump(
                            mode="json"
                        )
                        if case.execution_profile_key in execution_profiles
                        else None
                    ),
                    case_source="relational",
                )
            )
        return items

    if suite_id is not None:
        requested_suite_ids = [suite_id]
    else:
        requested_suite_ids = None

    stmt: Select[tuple[PLCTestSuiteRecord]] = select(PLCTestSuiteRecord)
    if suite_id is not None:
        stmt = stmt.where(PLCTestSuiteRecord.id == suite_id)
    suites = session.scalars(stmt.order_by(PLCTestSuiteRecord.created_at.desc())).all()
    if not suites:
        return []

    fallback_suite_ids = {suite.id for suite in suites}
    if requested_suite_ids is not None:
        fallback_suite_ids = set(requested_suite_ids)

    suites_with_relational_rows = set(
        session.scalars(
            select(PLCTestCaseRecord.suite_id)
            .where(PLCTestCaseRecord.suite_id.in_(fallback_suite_ids))
            .where(PLCTestCaseRecord.is_active.is_(True))
        ).all()
    )

    fallback_suites = [
        suite for suite in suites if suite.id not in suites_with_relational_rows
    ]
    if not fallback_suites:
        return []

    items: list[dict[str, Any]] = []
    for suite in fallback_suites:
        definition = PLCTestSuiteDefinitionModel.model_validate(suite.definition_json)
        for case in definition.cases:
            if instruction_name and case.instruction_name != instruction_name:
                continue
            if input_type and case.input_type != input_type:
                continue
            if tag and tag not in case.tags:
                continue
            items.append(
                {
                    "id": case.id,
                    "suite_id": suite.id,
                    "suite_title": suite.title,
                    "case_source": "definition_json_fallback",
                    **case.model_dump(mode="json"),
                }
            )
    return items


def create_plc_job_payload(
    session: Session,
    *,
    suite_id: str | None,
    testcase_ids: list[str] | None,
    target_key: str,
) -> tuple[PLCTestSuiteDetailModel, dict[str, Any], list[PLCTestCaseModel]]:
    if suite_id is None and not testcase_ids:
        raise PLCImportError("suite_id or testcase_ids is required")
    if suite_id is None and testcase_ids:
        suite_id = testcase_ids[0].split("::", 1)[0]
    suite = get_plc_suite_detail(session, suite_id or "")
    if suite is None:
        raise PLCImportError("PLC suite not found")

    selected_cases: list[PLCTestCaseModel]
    relational_cases = get_testcase_models_for_suite(session, suite_id=suite.id)
    if relational_cases:
        if len(relational_cases) != suite.case_count:
            raise PLCImportError(
                "PLC suite testcase masters are incomplete; reconcile relational rows before queueing a run"
            )
        selected_cases = relational_cases
    else:
        selected_cases = suite.definition_json.cases

    if testcase_ids:
        selected_case_set = set(testcase_ids)
        selected_cases = [
            case for case in selected_cases if case.id in selected_case_set
        ]
        if not selected_cases:
            raise PLCImportError("No PLC testcases matched testcase_ids")
    payload = {
        "suite_id": suite.id,
        "suite_title": suite.title,
        "target_key": target_key,
        "testcase_source": (
            "relational" if relational_cases else "definition_json_fallback"
        ),
        "testcases": [case.model_dump(mode="json") for case in selected_cases],
    }
    return suite, payload, selected_cases


def build_normalization_suggestion(raw_row: dict[str, Any]) -> dict[str, Any]:
    cases, warnings = _normalize_import_rows([raw_row])
    suggestion = _suite_definition_with_ids("suggestion", cases, warnings)
    return {
        "payload_schema_version": "plc-llm-suggestion.v1",
        "source_of_truth": False,
        "review_required": True,
        "suggestion_type": "normalization",
        "normalized_cases": [case.model_dump(mode="json") for case in suggestion.cases],
        "warnings": suggestion.warnings,
    }
